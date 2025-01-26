from products.models import Store, Product, Brand, StoreProduct
import xlrd
from decimal import Decimal
from tqdm import tqdm
from decimal import Decimal
from tenants.models import Tenant
from clients.models import Discount, Client
from openpyxl import Workbook
from django.conf import settings
import os


class TenantManager:
    """Clase para gestionar la creación de tiendas y usuarios."""

    def __init__(self, data_tenant):
        self.data_tenant = data_tenant

    def create_tenant(self):
        tenant, created = Tenant.objects.get_or_create(**self.data_tenant)
        return tenant


class StoreManager:
    """Clase para gestionar la creación de tiendas y usuarios."""

    def __init__(self, tenant):
        self.tenant = tenant

    def create_store(self, data):
        Store.objects.get_or_create(**data, tenant=self.tenant)

    def create_stores(self, stores_data):
        for data in stores_data:
            self.create_store(data)


class ProductManager:
    def __init__(self, tenant):
        self.tenant = tenant

    # Commons
    def read_excel_file(self, file_path):
        workbook = xlrd.open_workbook(file_path)
        return workbook.sheet_by_index(0)

    def clean_price(self, value):
        return Decimal(value.replace("$", "").replace(",", ""))

    # Demo
    def create_demo_product(self, demo_product):
        # Procesar marca
        code = demo_product.pop("code")
        brand_name = demo_product.pop("brand")
        # Crear objetos relacionados
        brand, _ = Brand.objects.get_or_create(name=brand_name, tenant=self.tenant)
        # Crear el producto
        product, _ = Product.objects.get_or_create(
            code=code, defaults={**demo_product, "brand": brand}
        )

        StoreProduct.objects.filter(product=product).update(stock=10)

    def create_demo_products(self, demo_products):
        for demo_product in tqdm(
            demo_products, desc="Creating Products", unit="product"
        ):
            self.create_demo_product(demo_product)

    # Eleventa
    def create_product_from_eleventa(self, row_values):
        code = str(row_values[0])
        name = row_values[1].lower().strip()

        purchase_price = self.clean_price(row_values[2])
        unit_sale_price = self.clean_price(row_values[3])
        wholesale_sale_price = (
            self.clean_price(row_values[4]) if row_values[4] else None
        )
        wholesale_sale_price = (
            wholesale_sale_price if unit_sale_price != wholesale_sale_price else None
        )
        min_wholesale_quantity = 3 if wholesale_sale_price else None

        # Procesar marca
        brand_name = (
            row_values[8]
            .replace(".", "")
            .replace(",", "")
            .replace("- Sin Departamento -", "NA")
            .strip()
            .title()
        )
        if len(brand_name) <= 2:
            brand_name = brand_name.upper()
        # Crear objetos relacionados
        brand, _ = Brand.objects.get_or_create(name=brand_name, tenant=self.tenant)

        # Datos del producto
        product_data = {
            "code": code,
            "name": name,
            "purchase_price": purchase_price,
            "unit_sale_price": unit_sale_price,
            "wholesale_sale_price": wholesale_sale_price,
            "min_wholesale_quantity": min_wholesale_quantity,
        }

        # Crear el producto
        Product.objects.get_or_create(
            code=code, defaults={**product_data, "brand": brand}
        )

    def is_product_in_tenant(self, code):
        return Product.objects.filter(code=code, brand__tenant=self.tenant).exists()

    def create_products_from_eleventa(self, file_path):
        file = self.read_excel_file(file_path)
        for row_idx in tqdm(
            range(1, file.nrows), desc="Creating Products", unit="product"
        ):
            row_values = file.row_values(row_idx)
            self.create_product_from_eleventa(row_values)

    def validate_products_in_tenant_from_eleventa(self, file_path):
        """Valida productos de un archivo Excel y guarda los inexistentes en un nuevo archivo."""
        # Leer el archivo de entrada con xlrd
        file = self.read_excel_file(file_path)

        # Crear un nuevo archivo Excel para productos inexistentes
        wb = Workbook()
        ws = wb.active  # Hoja activa
        ws.title = "Productos Inexistentes"

        # Escribir los nombres de las columnas (headers)
        headers = file.row_values(0)  # Primera fila como headers
        for col_index, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_index, value=header)

        # Inicializar el índice de fila para los datos (empezando después del header)
        row_index = 2

        # Iterar sobre las filas del archivo de entrada
        for row_idx in tqdm(
            range(1, file.nrows), desc="Checking Products", unit="product"
        ):
            row_values = file.row_values(row_idx)

            # Si el producto no está en el tenant, lo escribimos en el nuevo archivo
            if not self.is_product_in_tenant(row_values[0]):
                print(row_values)  # Opcional: Verificar en consola
                for col_index, value in enumerate(row_values, start=1):
                    ws.cell(row=row_index, column=col_index, value=value)
                row_index += 1

        # Guardar el nuevo archivo en la carpeta scripts
        output_file_path = os.path.join(
            settings.BASE_DIR,
            "scripts",
            f"Productos inexistentes en {self.tenant.name}.xlsx",
        )
        os.makedirs(
            os.path.dirname(output_file_path), exist_ok=True
        )  # Crear la carpeta si no existe
        wb.save(output_file_path)

        print(f"Archivo XLSX creado exitosamente en: {output_file_path}")

    def update_stock_by_store_from_eleventa(self, file_path, tenant, store_data):
        file = self.read_excel_file(file_path)
        store = Store.objects.get(tenant=tenant, **store_data)
        for row_idx in tqdm(
            range(1, file.nrows), desc="Updating prices", unit="product"
        ):
            row_values = file.row_values(row_idx)
            stock = self.clean_price(row_values[5])

            if stock == 0:
                continue

            code = row_values[0]

            product = Product.objects.get(code=code, brand__tenant=tenant)

            store_product = StoreProduct.objects.get(product=product, store=store)

            store_product.stock = stock
            store_product.save()

    def update_stock_by_store_from_smartventa(self, file_path, tenant, store_data):
        file = self.read_excel_file(file_path)
        store = Store.objects.get(tenant=tenant, **store_data)
        for row_idx in tqdm(
            range(1, file.nrows), desc="Updating prices", unit="product"
        ):
            row_values = file.row_values(row_idx)
            stock = row_values[2]
            if stock == 0:
                continue

            try:
                product = Product.objects.get(code=row_values[0], brand__tenant=tenant)
            except Product.DoesNotExist:
                code = row_values[0]
                brand_name = row_values[1].split()[0].capitalize()

                brand = Brand.objects.get(name=brand_name, tenant=tenant)
                name = " ".join(row_values[1].split()[1:])
                purchase_price = 0
                unit_sale_price = 0
                data = {
                    "code": code,
                    "brand": brand,
                    "name": name,
                    "purchase_price": purchase_price,
                    "unit_sale_price": unit_sale_price,
                }
                product = Product.objects.create(**data)

            store_product = StoreProduct.objects.get(product=product, store=store)

            store_product.stock = stock
            store_product.save()


class ClientManager:
    def __init__(self, tenant):
        self.tenant = tenant

    def create_demo_client(self, demo_client):
        # Procesar marca
        discount = demo_client.pop("discount_percentage")
        discount, _ = Discount.objects.get_or_create(
            discount_percentage=discount, tenant=self.tenant
        )
        # Crear el producto
        client, _ = Client.objects.get_or_create(
            defaults={**demo_client, "discount": discount}
        )

    def create_demo_clients(self, demo_clients):
        for demo_client in tqdm(demo_clients, desc="Creating Products", unit="product"):
            self.create_demo_client(demo_client)
